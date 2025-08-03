import kivy
from kivy.app import App
from kivy.uix.gridlayout import GridLayout
from kivy.properties import ObjectProperty
import os, sys
from kivy.resources import resource_add_path
from kivy.lang import Builder
import openpyxl
from openpyxl import Workbook, load_workbook


class MyLayout(GridLayout):

    gray = (0.4,0.4,0.4,1)
    red = (245/255.0, 115/255.0, 115/255.0,1)
    blue = (115/255.0, 182/255.0, 245/255.0,1)
    green = (164/255.0, 250/255.0, 132/255.0,1)
    purple = (212/255.0, 174/255.0, 242/255.0,1)
    yellow = (240/255.0, 235/255.0, 144/255.0,1)

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    xl = load_workbook('checklist.xlsx')
    sheet = xl['Sheet1']
    
    cells = ['A1', 'A2', 'A3', 'A4', 'A5']
    check_cells = ['B1', 'B2', 'B3', 'B4', 'B5']

    for cell in cells:
        if sheet[cell].value is None:
            sheet[cell].value = " "

    is_checked = [False, False, False, False, False]

    
    if sheet[check_cells[0]].value == "True":
        is_checked[0] = True
    if sheet[check_cells[1]].value == "True":
        is_checked[1] = True
    if sheet[check_cells[2]].value == "True":
        is_checked[2] = True
    if sheet[check_cells[3]].value == "True":
        is_checked[3] = True
    if sheet[check_cells[4]].value == "True":
        is_checked[4] = True

    text1 = sheet['A1']
    text2 = sheet['A2']
    text3 = sheet['A3']
    text4 = sheet['A4']
    text5 = sheet['A5']

    def day_selected(self, value):
        if value == "Monday":
            cells = ['A1', 'A2', 'A3', 'A4', 'A5']
            self.check_cells = ['B1', 'B2', 'B3', 'B4', 'B5']
        elif value == "Tuesday":
            cells = ['C1', 'C2', 'C3', 'C4', 'C5']
            self.check_cells = ['D1', 'D2', 'D3', 'D4', 'D5']
        elif value == "Wednesday":
            cells = ['E1', 'E2', 'E3', 'E4', 'E5']
            self.check_cells = ['F1', 'F2', 'F3', 'F4', 'F5']
        elif value == "Thursday":
            cells = ['G1', 'G2', 'G3', 'G4', 'G5']
            self.check_cells = ['H1', 'H2', 'H3', 'H4', 'H5']
        elif value == "Friday":
            cells = ['I1', 'I2', 'I3', 'I4', 'I5']
            self.check_cells = ['J1', 'J2', 'J3', 'J4', 'J5']
        elif value == "Saturday":
            cells = ['K1', 'K2', 'K3', 'K4', 'K5']
            self.check_cells = ['L1', 'L2', 'L3', 'L4', 'L5']
        elif value == "Sunday":
            cells = ['M1', 'M2', 'M3', 'M4', 'M5']
            self.check_cells = ['N1', 'N2', 'N3', 'N4', 'N5']


        for cell in cells:
            if self.sheet[cell].value is None:
                    self.sheet[cell].value = " "

        self.text1 = self.sheet[cells[0]]
        self.text2 = self.sheet[cells[1]]
        self.text3 = self.sheet[cells[2]]
        self.text4 = self.sheet[cells[3]]
        self.text5 = self.sheet[cells[4]]
        self.ids.task1.text = self.text1.value
        self.ids.task2.text = self.text2.value
        self.ids.task3.text = self.text3.value
        self.ids.task4.text = self.text4.value
        self.ids.task5.text = self.text5.value


        if self.sheet[self.check_cells[0]].value == "True":
            self.ids.check1.active = True
        else:
            self.ids.check1.active = False
            
        if self.sheet[self.check_cells[1]].value == "True":
            self.ids.check2.active = True
        else:
            self.ids.check2.active = False
            
        if self.sheet[self.check_cells[2]].value == "True":
            self.ids.check3.active = True
        else:
            self.ids.check3.active = False
            
        if self.sheet[self.check_cells[3]].value == "True":
            self.ids.check4.active = True
        else:
            self.ids.check4.active = False
            
        if self.sheet[self.check_cells[4]].value == "True":
            self.ids.check5.active = True
        else:
            self.ids.check5.active = False

    def checked(self, instance, value, cell, which_box, which_color):
        if value:
            self.sheet[cell].value = "True"
            which_box.background_color = self.gray
        else:
            self.sheet[cell].value = "False"
            which_box.background_color = which_color
            
        self.xl.save('checklist.xlsx')
        
    def save_data(self):
        self.text1.value = self.ids.task1.text
        self.text2.value = self.ids.task2.text
        self.text3.value = self.ids.task3.text
        self.text4.value = self.ids.task4.text
        self.text5.value = self.ids.task5.text
        self.xl.save('checklist.xlsx')
        
    def clear_tasks(self):
        self.ids.task1.text = ""
        self.ids.task2.text = ""
        self.ids.task3.text = ""
        self.ids.task4.text = ""
        self.ids.task5.text = ""
        self.save_data()
    def clear_checks(self):
        self.ids.check1.active = False
        self.ids.check2.active = False
        self.ids.check3.active = False
        self.ids.check4.active = False
        self.ids.check5.active = False

    
class ChecklistApp(App):
    def build(self):
        self.icon = self.resource_path('check.png')
        xl2 = load_workbook(self.resource_path('checklist.xlsx'))
        return MyLayout()
        return Builder.load_file(self.resource_path('checklist.kv'))
    def start_app(self):
        self.ids.task1.text = self.sheet['A1']
        
    @staticmethod
    def resource_path(relative_path):

        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath('.')
        return os.path.join(base_path, relative_path)



if __name__ == '__main__':

    if hasattr(sys, '_MEIPASS'):
        resource_add_path((os.path.join(sys._MEIPASS)))
    
    ChecklistApp().run()


